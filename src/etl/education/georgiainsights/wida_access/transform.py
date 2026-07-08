"""Transform bronze wida_access files into gold fact tables.

Source: Georgia Insights (GaDOE) WIDA ACCESS for ELLs state-level results —
one ``.xlsx`` per spring testing year, 2017-2024 (no file for 2016 or
earlier). Each file's single ``State`` sheet is a wide pivot-style report:
for every grade (K-12), the count and share of Georgia English Learners
scoring at each of the six WIDA proficiency levels (Level 1 Entering ...
Level 6 Reaching) in each of eight domains — four language domains
(Listening, Speaking, Reading, Writing) and four composites (Oral Language,
Literacy, Comprehension, Overall Score).

Design decisions (every invariant re-verified against THIS topic's 8 bronze
files during authoring — see bronze-data-structure.md):

- **Multi-row pivot headers require openpyxl.** ``pl.read_excel`` reads only
  row 1 (the title) as the header and returns ``__UNNAMED__N`` columns. The
  local reader pulls the sheet with openpyxl and rebuilds one flat header
  per column as ``{domain} | {level} | {metric}`` by forward-filling the
  merged-cell ``None`` gaps in the domain row (row 2) and level row (row 3),
  then attaching the metric row (row 4). Whitespace (including embedded
  newlines like ``Level 1\\nEntering``) collapses to single spaces. Excel
  reads load whole sheets, so read-loss raw == parsed by construction.

- **Two eras, detected by column signature.** Era 2 is the 2021 COVID file
  only: it adds per-domain ``Total Tested in Domain`` denominators and
  enrollment-based testing-participation rates (testing was voluntary /
  disrupted that year), and its sheet carries 8 trailing footnote/blank rows
  that are filtered out (kept rows must have a literal grade token ``K`` /
  ``1``..``12`` in column 0; the filter is ledgered via
  ``manifest.record_filtered``). Era 1 (the other 7 files) publishes neither
  the denominators nor the participation rates.

- **Era-2 composite labels carry footnote letters.** 2021's composite
  domain headers are ``Oral Language CompositeA`` / ``Literacy CompositeB``
  / ``Comprehension CompositeC`` / ``Overall Score CompositeD`` — the
  trailing letter references the composite-weighting footnotes in the
  filtered rows. Rather than regex-stripping, ``DOMAIN_MAP`` enumerates all
  12 observed bronze labels (8 base + 4 footnote variants) so the manifest
  records the labels actually seen in each era.

- **`pct_at_proficiency_level`'s denominator is the per-domain total, NOT
  `num_tested`.** Verified in bronze: the six level shares sum to 100.0 per
  (grade, domain) within 2e-06 in every file, and the six level counts sum
  to a per-domain total that is <= the overall ``Total Number of Students
  Tested`` (not every student tests in every domain). Era 2 publishes that
  denominator explicitly (``Total Tested in Domain``, which equals the
  level-count sum exactly — 0 mismatches in bronze); for Era 1 the transform
  reconstructs ``num_tested_in_domain`` as the sum of the six level counts
  so the denominator is populated in every year.

- **Grade is the row axis** (no race/gender/economic breakouts), so it lives
  in ``grade_level`` (canonical codes ``k``, ``01``..``12`` via the shared
  ``src/utils/grades`` normalizer) and there is no ``demographic`` column,
  per the education domain grade-in-demographic policy.

- **State detail level only.** Bronze publishes no district or school
  breakdown; ``district_code`` and ``school_code`` are NULL on every row and
  only ``states.parquet`` is emitted per year.

- **Percent columns are 0-100 in bronze** (verified: level shares sum to
  100 per domain; enrollment-participation max is 95.45) and are divided by
  100 to the canonical 0-1 scale. All are bounded proportions (max single
  level share observed is 80.8) — ``unit: proportion``.

- **No suppression.** Zero suppression markers in any file (state-level
  aggregates exceed any small-cell threshold); zero NULLs in bronze. A ``0``
  count is a real zero. ``suppressed_to_null=False`` on the contract.

- **No §4b masks.** Every observed value is within its metric's defined
  scale (counts >= 0; proportions within [0, 1] after scaling).

- **Dedup tie-break**: one file per year with a unique (grade, domain,
  level) grain inside each file, so no natural-key duplicates are expected;
  ``sort_col="num_at_proficiency_level"`` is the documented defensive
  tie-break (prefer the row carrying a reported, larger count over a
  placeholder). The collision guard runs first and raises on any
  divergent-metric duplicate.
"""

import logging
import re
from pathlib import Path

import openpyxl
import polars as pl

from src.utils.grades import GRADE_LEVEL_MAP, normalize_grade_column
from src.utils.metadata import write_data_dictionary
from src.utils.readers import list_bronze_files
from src.utils.transformers import (
    TransformManifest,
    assert_no_natural_key_collisions,
    deduplicate_by_detail_level,
    detect_era_by_columns,
    export_to_parquet,
    harmonize_columns,
    null_aggregate_geography,
    validate_output,
)
from src.utils.validators import (
    EDUCATION_DOMAIN_CONFIG,
    check_null_rate_spikes,
    run_topic_validation,
)

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(name)s - %(levelname)s - %(message)s",
)
logger = logging.getLogger(__name__)


# =============================================================================
# Configuration
# =============================================================================

TOPIC = "wida_access"
BRONZE_DIR = Path("data/bronze/education/georgiainsights/wida_access")
GOLD_DIR = Path("data/gold/education/wida_access")
SOURCE_URL = "https://georgiainsights.gadoe.org/data-downloads/"

# Every bronze workbook has exactly one sheet, named `State` (verified).
SHEET_NAME = "State"

# 0-indexed sheet rows: row 0 is the title string, rows 1-3 are the
# domain / level / metric header levels, data rows start at row 4.
DOMAIN_ROW_IDX = 1
LEVEL_ROW_IDX = 2
METRIC_ROW_IDX = 3
DATA_START_ROW_IDX = 4

# Three filename patterns exist (underscores 2017-2020/2022-2023, spaces
# 2021, hyphens 2024); each contains exactly one 4-digit 20YY year token.
FILENAME_YEAR_PATTERN = re.compile(r"(20\d{2})")

# Valid grade tokens in column 0 of a data row. Everything else in the data
# region (blank separators, 2021's composite-weighting footnotes and URL) is
# filtered out and ledgered.
VALID_GRADE_TOKENS: set[str] = {"K"} | {str(g) for g in range(1, 13)}

# Bronze domain label -> gold domain code. The four Era-2 (2021) composite
# labels carry a trailing footnote-reference letter (A-D); enumerating them
# beside the Era-1 base labels keeps the manifest's bronze_values_seen
# faithful to what each file actually publishes (no regex stripping).
DOMAIN_MAP: dict[str, str] = {
    "Listening Domain": "listening",
    "Speaking Domain": "speaking",
    "Reading Domain": "reading",
    "Writing Domain": "writing",
    "Oral Language Composite": "oral_language_composite",
    "Oral Language CompositeA": "oral_language_composite",
    "Literacy Composite": "literacy_composite",
    "Literacy CompositeB": "literacy_composite",
    "Comprehension Composite": "comprehension_composite",
    "Comprehension CompositeC": "comprehension_composite",
    "Overall Score Composite": "overall_score_composite",
    "Overall Score CompositeD": "overall_score_composite",
}

# WIDA proficiency-level label -> gold code. Keys are the
# whitespace-normalized form of the bronze level-row labels (bronze embeds a
# newline: `Level 1\nEntering`); the WIDA six-level scale is fixed and
# identical in every file.
PROFICIENCY_LEVEL_MAP: dict[str, str] = {
    "Level 1 Entering": "level_1_entering",
    "Level 2 Emerging": "level_2_emerging",
    "Level 3 Developing": "level_3_developing",
    "Level 4 Expanding": "level_4_expanding",
    "Level 5 Bridging": "level_5_bridging",
    "Level 6 Reaching": "level_6_reaching",
}

# Per-era flat header of the overall (at-least-one-domain) tested count.
# Era 2 renamed the column to make the "at least one domain" semantic
# explicit; both mean the same thing.
TOTAL_TESTED_HEADER: dict[str, str] = {
    "era_1_standard": "Total Number of Students Tested",
    "era_2_2021_participation": (
        "Total Number of Students Tested in At Least One Domain"
    ),
}

# Era-2-only flat headers (overall and per-domain participation metrics).
OVERALL_PCT_ENROLLED_HEADER = (
    "Percentage of Enrolled Students Tested in At Least One Domain"
)
PCT_ENROLLED_IN_DOMAIN_PREFIX = "Percentage of Enrolled Students Tested in "

# Gold fact column order. `detail_level` is carried through dedup /
# geography-nulling / export-splitting, then dropped by export_to_parquet().
STANDARD_COLUMNS: list[str] = [
    "year",
    "district_code",
    "school_code",
    "grade_level",
    "domain",
    "proficiency_level",
    "num_at_proficiency_level",
    "pct_at_proficiency_level",
    "num_tested_in_domain",
    "num_tested",
    "enrolled_tested_rate",
    "enrolled_tested_in_domain_rate",
    "detail_level",
]

TARGET_TYPES: dict[str, pl.DataType] = {
    "year": pl.Int32,
    "district_code": pl.Utf8,
    "school_code": pl.Utf8,
    "grade_level": pl.Utf8,
    "domain": pl.Utf8,
    "proficiency_level": pl.Utf8,
    "num_at_proficiency_level": pl.Int64,
    "pct_at_proficiency_level": pl.Float64,
    "num_tested_in_domain": pl.Int64,
    "num_tested": pl.Int64,
    "enrolled_tested_rate": pl.Float64,
    "enrolled_tested_in_domain_rate": pl.Float64,
    "detail_level": pl.Utf8,
}

METRIC_COLUMNS: list[str] = [
    "num_at_proficiency_level",
    "pct_at_proficiency_level",
    "num_tested_in_domain",
    "num_tested",
    "enrolled_tested_rate",
    "enrolled_tested_in_domain_rate",
]

NATURAL_KEYS: list[str] = [
    "year",
    "district_code",
    "school_code",
    "detail_level",
    "grade_level",
    "domain",
    "proficiency_level",
]


# =============================================================================
# Bronze reading — multi-row pivot header flattening
# =============================================================================


def _normalize_whitespace(value: object) -> str | None:
    """Collapse runs of whitespace (incl. embedded newlines) to one space.

    Bronze headers embed literal newlines (`Level 1\\nEntering`,
    `Total \\nNumber of Students Tested`); normalizing at read time gives
    every downstream lookup one predictable canonical form.
    """
    if value is None:
        return None
    return re.sub(r"\s+", " ", str(value)).strip()


def _forward_fill(values: list[object]) -> list[str | None]:
    """Forward-fill a header row's merged-cell ``None`` gaps.

    openpyxl reports a merged range's value only in its first cell; the rest
    are ``None``. Filling forward reconstructs the per-column domain / level
    label. Leading ``None`` cells stay ``None``.
    """
    filled: list[str | None] = []
    last: str | None = None
    for v in values:
        normalized = _normalize_whitespace(v)
        if normalized is not None:
            last = normalized
        filled.append(last)
    return filled


def _read_state_sheet(path: Path) -> tuple[list[list[object]], list[str]]:
    """Read the ``State`` sheet and build one flat header per column.

    Returns ``(rows, flat_headers)`` where ``rows`` is the full raw sheet
    (lists, padded to a uniform width) and ``flat_headers[i]`` is the
    ``{domain} | {level} | {metric}`` join of the non-``None`` header levels
    for column ``i`` (e.g. plain ``Grade`` for the row-axis column).
    """
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        if SHEET_NAME not in wb.sheetnames:
            raise ValueError(
                f"{path.name}: expected sheet {SHEET_NAME!r}, found {wb.sheetnames}"
            )
        rows = [list(r) for r in wb[SHEET_NAME].iter_rows(values_only=True)]
    finally:
        wb.close()

    if len(rows) <= DATA_START_ROW_IDX:
        raise ValueError(
            f"{path.name}: sheet has only {len(rows)} rows; expected headers "
            f"in rows 1-{DATA_START_ROW_IDX} plus data rows"
        )

    # Pad ragged rows to a uniform width so per-column zips don't drop
    # trailing columns when openpyxl reports a sparse row as shorter.
    ncols = max(len(r) for r in rows)
    for r in rows:
        r.extend([None] * (ncols - len(r)))

    domain_filled = _forward_fill(rows[DOMAIN_ROW_IDX])
    level_filled = _forward_fill(rows[LEVEL_ROW_IDX])
    metric_row = [_normalize_whitespace(v) for v in rows[METRIC_ROW_IDX]]

    flat_headers: list[str] = []
    for idx, (dom, lvl, met) in enumerate(
        zip(domain_filled, level_filled, metric_row, strict=True)
    ):
        parts = [p for p in (dom, lvl, met) if p is not None]
        # A fully blank header column would silently collapse into a
        # duplicate name; tag it uniquely so the duplicate-header guard
        # below sees it. (Not observed in any bronze file.)
        flat_headers.append(" | ".join(parts) if parts else f"__BLANK_COL_{idx}__")

    if len(set(flat_headers)) != len(flat_headers):
        dupes = sorted({h for h in flat_headers if flat_headers.count(h) > 1})
        raise ValueError(f"{path.name}: duplicate flattened headers: {dupes}")

    return rows, flat_headers


def _check_title_year(rows: list[list[object]], year: int, filename: str) -> None:
    """Cross-check the row-1 title's embedded year against the filename year.

    Every file's title row carries `Spring YYYY ... State Results` (cell
    position varies). A mismatch means a misnamed or mis-filled file — fail
    loudly rather than label a year wrong.
    """
    for cell in rows[0]:
        if cell is None:
            continue
        match = FILENAME_YEAR_PATTERN.search(str(cell))
        if match and int(match.group(1)) != year:
            raise ValueError(
                f"{filename}: title row says year {match.group(1)} but the "
                f"filename year is {year}"
            )


def _extract_grade_rows(
    rows: list[list[object]], headers: list[str]
) -> tuple[pl.DataFrame, int, list[str]]:
    """Extract the grade data rows into an all-Utf8 DataFrame.

    Keeps a data-region row iff its column-0 token is ``K`` or ``1``..``12``
    — this drops the 2021 file's blank separators, composite-weighting
    footnotes, and URL row. Every cell is stringified so Polars never has to
    infer the mixed int/str Grade column (``K`` is a string; 1-12 are ints)
    or the mixed int/float metric cells; numeric casts happen downstream
    with explicit ``strict=False``.

    Returns ``(df, data_region_rows, dropped_samples)`` where
    ``data_region_rows`` counts all rows below the headers and
    ``dropped_samples`` holds the non-grade tokens removed (for the filter
    ledger).
    """
    data_rows: list[list[str | None]] = []
    dropped_samples: list[str] = []
    data_region = rows[DATA_START_ROW_IDX:]
    for r in data_region:
        token = None if r[0] is None else str(r[0]).strip()
        if token in VALID_GRADE_TOKENS:
            data_rows.append([None if v is None else str(v) for v in r])
        else:
            dropped_samples.append("<blank>" if token is None else token[:60])

    if not data_rows:
        return (
            pl.DataFrame(schema={h: pl.Utf8 for h in headers}),
            len(data_region),
            dropped_samples,
        )
    df = pl.DataFrame(data_rows, schema=headers, orient="row")
    return df, len(data_region), dropped_samples


# =============================================================================
# Era transform
# =============================================================================


def _require_column(df: pl.DataFrame, column: str, label: str) -> None:
    """Raise if an expected flattened bronze column is absent.

    An unmatched source column silently becomes NULL in gold (the most
    common data-loss bug), so a missing column fails loudly instead.
    """
    if column not in df.columns:
        raise ValueError(
            f"{label}: expected bronze column {column!r} not found. "
            f"First columns present: {df.columns[:6]}"
        )


def _resolve_domain_labels(df: pl.DataFrame, label: str) -> dict[str, str]:
    """Map each gold domain to the bronze label this file uses for it.

    A file uses either the Era-1 base label (`Oral Language Composite`) or
    the Era-2 footnote variant (`Oral Language CompositeA`) — never both.
    Raises when a gold domain matches zero or multiple bronze labels.
    """
    resolved: dict[str, str] = {}
    for bronze_label, gold_domain in DOMAIN_MAP.items():
        # A domain's block always contains the Level 1 count column; its
        # presence is the per-file signature for that bronze label.
        marker = f"{bronze_label} | Level 1 Entering | # of Students at Level"
        if marker not in df.columns:
            continue
        if gold_domain in resolved:
            raise ValueError(
                f"{label}: gold domain {gold_domain!r} matches multiple "
                f"bronze labels ({resolved[gold_domain]!r} and {bronze_label!r})"
            )
        resolved[gold_domain] = bronze_label
    missing = sorted(set(DOMAIN_MAP.values()) - set(resolved))
    if missing:
        raise ValueError(f"{label}: no bronze columns found for domain(s) {missing}")
    return resolved


def _transform_era(
    df: pl.DataFrame, year: int, era: str, manifest: TransformManifest
) -> pl.DataFrame:
    """Unpivot one year's flattened sheet into the long-form fact table.

    Both eras share the (domain x level x metric) pivot core, so one loop
    handles either; Era 2 additionally contributes the published per-domain
    denominators and the enrollment-participation rates (NULL in Era 1).
    Output: one row per (grade, domain, proficiency_level) — 13 x 8 x 6 =
    624 rows per year.
    """
    label = f"{TOPIC} {year}"
    is_era_2 = era == "era_2_2021_participation"
    grade_col = "Grade"
    total_col = TOTAL_TESTED_HEADER[era]
    _require_column(df, grade_col, label)
    _require_column(df, total_col, label)

    # --- Per-grade base frame (repeated onto every domain x level row) ---
    # Counts hop through Float64 so a float-formatted count string would
    # survive the cast (bronze counts are int-typed in every file — the hop
    # is defensive and value-identical for clean integers).
    base = df.select(
        pl.lit(year).cast(pl.Int32).alias("year"),
        pl.lit(None).cast(pl.Utf8).alias("district_code"),
        pl.lit(None).cast(pl.Utf8).alias("school_code"),
        pl.lit("state").alias("detail_level"),
        normalize_grade_column(grade_col).alias("grade_level"),
        pl.col(grade_col).str.strip_chars().alias("_grade_raw"),
        pl.col(total_col)
        .cast(pl.Float64, strict=False)
        .cast(pl.Int64, strict=False)
        .alias("num_tested"),
    )

    # Record the effective slice of the shared grade map — the spellings this
    # file actually hit — so map_used stays reviewable while the unmapped
    # guard still applies.
    observed_grades = df[grade_col].drop_nulls().unique().to_list()
    effective_grade_map = {
        g: GRADE_LEVEL_MAP[g.strip().upper()]
        for g in observed_grades
        if g.strip().upper() in GRADE_LEVEL_MAP
    }
    manifest.record_categorical(
        column="grade_level",
        map_dict=effective_grade_map,
        bronze_series=df[grade_col],
        gold_series=base["grade_level"],
    )

    # Era 2 publishes the overall participation rate; Era 1 never does —
    # synthesize NULL so every year carries the same column set. 0-100 ->
    # 0-1 per data-cleaning-standards §4.
    if is_era_2:
        _require_column(df, OVERALL_PCT_ENROLLED_HEADER, label)
        base = base.with_columns(
            (
                df[OVERALL_PCT_ENROLLED_HEADER].cast(pl.Float64, strict=False) / 100.0
            ).alias("enrolled_tested_rate")
        )
    else:
        base = base.with_columns(
            pl.lit(None).cast(pl.Float64).alias("enrolled_tested_rate")
        )

    # --- Unpivot the 8 domains x 6 levels x 2 metrics core ---------------
    domain_labels = _resolve_domain_labels(df, label)
    domain_frames: list[pl.DataFrame] = []
    for gold_domain, bronze_label in domain_labels.items():
        level_frames: list[pl.DataFrame] = []
        count_cols: list[str] = []
        for bronze_level, gold_level in PROFICIENCY_LEVEL_MAP.items():
            count_col = f"{bronze_label} | {bronze_level} | # of Students at Level"
            pct_col = f"{bronze_label} | {bronze_level} | % of Total Tested"
            _require_column(df, count_col, label)
            _require_column(df, pct_col, label)
            count_cols.append(count_col)
            level_frames.append(
                df.select(
                    pl.col(grade_col).str.strip_chars().alias("_grade_raw"),
                    pl.lit(gold_domain).alias("domain"),
                    pl.lit(gold_level).alias("proficiency_level"),
                    pl.col(count_col)
                    .cast(pl.Float64, strict=False)
                    .cast(pl.Int64, strict=False)
                    .alias("num_at_proficiency_level"),
                    # 0-100 -> 0-1 scale per data-cleaning-standards §4.
                    (pl.col(pct_col).cast(pl.Float64, strict=False) / 100.0).alias(
                        "pct_at_proficiency_level"
                    ),
                )
            )
        domain_long = pl.concat(level_frames, how="vertical")

        # Per-domain denominator: Era 2 publishes it (`Total Tested in
        # Domain`, verified equal to the level-count sum in every 2021
        # cell — the published value is preferred as authoritative); Era 1
        # reconstructs it as the sum of the six level counts (verified in
        # bronze: the six shares sum to 100.0 against exactly this total).
        if is_era_2:
            denom_col = f"{bronze_label} | Total Tested in Domain"
            _require_column(df, denom_col, label)
            denom_expr = pl.col(denom_col).cast(pl.Float64, strict=False).cast(pl.Int64)
        else:
            denom_expr = pl.sum_horizontal(
                *[pl.col(c).cast(pl.Int64, strict=False) for c in count_cols]
            )
        per_grade = df.select(
            pl.col(grade_col).str.strip_chars().alias("_grade_raw"),
            denom_expr.alias("num_tested_in_domain"),
        )

        # Per-domain participation rate (Era 2 only). The header tail names
        # the domain count the composite spans: `in Domain` (the 4 language
        # domains), `in Both Domains` (2-domain composites), `in All Four
        # Domains` (Overall Score) — matched by shared prefix.
        if is_era_2:
            pct_enrolled_cols = [
                c
                for c in df.columns
                if c.startswith(f"{bronze_label} | {PCT_ENROLLED_IN_DOMAIN_PREFIX}")
            ]
            if len(pct_enrolled_cols) != 1:
                raise ValueError(
                    f"{label}: expected exactly one enrolled-participation "
                    f"column for {bronze_label!r}, found {pct_enrolled_cols}"
                )
            per_grade = per_grade.with_columns(
                (df[pct_enrolled_cols[0]].cast(pl.Float64, strict=False) / 100.0).alias(
                    "enrolled_tested_in_domain_rate"
                )
            )
        else:
            per_grade = per_grade.with_columns(
                pl.lit(None).cast(pl.Float64).alias("enrolled_tested_in_domain_rate")
            )

        domain_frames.append(domain_long.join(per_grade, on="_grade_raw", how="left"))

    long_df = pl.concat(domain_frames, how="vertical")

    manifest.record_categorical(
        column="domain",
        map_dict=DOMAIN_MAP,
        bronze_series=pl.Series(sorted(domain_labels.values())),
        gold_series=long_df["domain"],
    )
    manifest.record_categorical(
        column="proficiency_level",
        map_dict=PROFICIENCY_LEVEL_MAP,
        # Bronze level labels after whitespace normalization (raw cells
        # embed a newline: `Level 1\nEntering`).
        bronze_series=pl.Series(list(PROFICIENCY_LEVEL_MAP)),
        gold_series=long_df["proficiency_level"],
    )

    result = long_df.join(base, on="_grade_raw", how="left")
    if result.height != df.height * len(domain_labels) * len(PROFICIENCY_LEVEL_MAP):
        raise ValueError(
            f"{label}: unpivot produced {result.height} rows; expected "
            f"{df.height} grades x {len(domain_labels)} domains x "
            f"{len(PROFICIENCY_LEVEL_MAP)} levels"
        )
    return result.select(STANDARD_COLUMNS)


# =============================================================================
# File dispatcher
# =============================================================================


def transform_file(path: Path, manifest: TransformManifest) -> pl.DataFrame | None:
    """Read one bronze workbook and transform it into long-form rows."""
    match = FILENAME_YEAR_PATTERN.search(path.name)
    if match is None:
        raise ValueError(f"Cannot extract year from filename: {path.name}")
    year = int(match.group(1))

    rows, headers = _read_state_sheet(path)
    _check_title_year(rows, year, path.name)
    df, data_region_rows, dropped_samples = _extract_grade_rows(rows, headers)

    # Whole-sheet Excel read: nothing can be dropped at parse time, so
    # raw == parsed by construction (record_read_loss is a no-op).
    manifest.record_read_loss(year, path.name, data_region_rows, data_region_rows)
    # Non-grade rows in the data region (2021 only: 2 blank separators, 4
    # composite-weighting footnotes, 1 note, 1 URL) are template furniture,
    # not observations — ledger the filter explicitly.
    if dropped_samples:
        logger.info(
            "Year %d: filtered %d non-grade data-region row(s): %s",
            year,
            len(dropped_samples),
            dropped_samples,
        )
        manifest.record_filtered(
            year, len(dropped_samples), "non_grade_footnote_or_blank_rows"
        )

    # Era detection by column signature: only the 2021 file publishes the
    # per-domain `Total Tested in Domain` denominator columns.
    era = detect_era_by_columns(
        df,
        {
            "era_2_2021_participation": ["Listening Domain | Total Tested in Domain"],
            "era_1_standard": [
                "Listening Domain | Level 1 Entering | # of Students at Level"
            ],
        },
    )
    if era is None:
        raise ValueError(
            f"{path.name}: could not detect era from flattened columns; "
            f"first columns: {df.columns[:6]}"
        )

    manifest.record_file(path, year, era, data_region_rows, df.columns)
    manifest.record_bronze(year, data_region_rows)

    if df.height == 0:
        logger.warning("Year %d: no grade rows in %s, skipping", year, path.name)
        return None
    logger.info(
        "Processing %s (year %d, %s): %d grade rows, %d columns",
        path.name,
        year,
        era,
        df.height,
        len(df.columns),
    )
    return _transform_era(df, year, era, manifest)


# =============================================================================
# Main pipeline
# =============================================================================


def main() -> None:
    """Run the full bronze-to-gold pipeline for wida_access."""
    manifest = TransformManifest(topic=TOPIC, bronze_dir=BRONZE_DIR, gold_dir=GOLD_DIR)

    # 1. Read + transform each bronze file (one file per year).
    all_dfs: list[pl.DataFrame] = []
    for path in list_bronze_files(BRONZE_DIR, extensions=[".xlsx"]):
        result = transform_file(path, manifest)
        if result is not None and result.height > 0:
            all_dfs.append(result)
    if not all_dfs:
        raise ValueError(f"No bronze data produced any rows under {BRONZE_DIR}")

    # 2. Harmonize columns/dtypes across eras and concatenate.
    all_dfs = harmonize_columns(all_dfs, STANDARD_COLUMNS, TARGET_TYPES)
    combined = pl.concat(all_dfs, how="vertical")
    logger.info("Combined %d rows across %d files", combined.height, len(all_dfs))

    # 3. Collision guard BEFORE dedup: duplicate keys with divergent metrics
    # mean a label-collapse bug and must raise, not be deduped away.
    assert_no_natural_key_collisions(
        combined, natural_keys=NATURAL_KEYS, metric_cols=METRIC_COLUMNS
    )
    # Tie-break: one workbook per year with a unique (grade, domain, level)
    # grain inside each file, so no duplicates are expected; prefer the row
    # with a reported (non-null, larger) count as the defensive safety net.
    combined = deduplicate_by_detail_level(
        combined,
        school_keys=[],  # no school rows in this topic
        district_keys=[],  # no district rows in this topic
        state_keys=["year", "grade_level", "domain", "proficiency_level"],
        sort_col="num_at_proficiency_level",
    )

    # 4. Geography nulling (shared domain rules; all rows are state-level so
    # both codes end NULL). No §4b masks: every observed value is within its
    # metric's defined scale (see module docstring).
    combined = null_aggregate_geography(
        combined,
        detail_level_col="detail_level",
        geography_rules=EDUCATION_DOMAIN_CONFIG["detail_level_geography_rules"],
    )

    # Pre-export sanity. The two participation-rate columns are published
    # only in 2021, so the spike check flags their 100%%-NULL Era-1 years —
    # a true, documented signal (not suppressed).
    spike_result = check_null_rate_spikes(combined, METRIC_COLUMNS)
    if spike_result.status == "warning":
        logger.warning("NULL-rate spikes: %s", spike_result.details)
    validate_output(
        combined,
        required_non_null=[
            "year",
            "detail_level",
            "grade_level",
            "domain",
            "proficiency_level",
        ],
    )

    # 5. Manifest stats on the FINAL DataFrame, then export.
    manifest.record_gold_from_dataframe(combined)
    manifest.compute_metric_stats(combined, METRIC_COLUMNS)
    export_to_parquet(combined, GOLD_DIR, STANDARD_COLUMNS)
    manifest.write(GOLD_DIR)

    # 6. Contract + README from the in-code column declaration.
    _emit_contract_and_readme(
        year_range=(int(combined["year"].min()), int(combined["year"].max()))
    )

    summary = manifest.tracker.summary()
    logger.info(
        "Done. Bronze rows: %s; gold rows: %s; years: %s",
        f"{summary['total_bronze']:,}",
        f"{summary['total_gold']:,}",
        summary["years_processed"],
    )

    # 7. ALWAYS LAST: validate the gold just written against the contract
    # just emitted. Raises GoldValidationError -> non-zero exit.
    run_topic_validation(GOLD_DIR)


def _emit_contract_and_readme(year_range: tuple[int, int]) -> None:
    """Emit the ODCS contract and gold README via ``write_data_dictionary``.

    The column declaration order MUST match STANDARD_COLUMNS minus
    ``detail_level`` — the contract properties (and the validator's schema
    check) follow it.
    """
    write_data_dictionary(
        output_dir=GOLD_DIR,
        name=TOPIC,
        description=(
            "Georgia Insights (GaDOE) WIDA ACCESS for ELLs state-level "
            "results — the spring English-language-proficiency assessment "
            "for Georgia English Learners. For each grade (kindergarten "
            "through 12) and each of eight WIDA domains — four language "
            "domains (listening, speaking, reading, writing) and four "
            "composites (oral language, literacy, comprehension, overall "
            "score) — reports the count and share of ELL students scoring "
            "at each of the six WIDA proficiency levels (Level 1 Entering "
            "through Level 6 Reaching). This is a proficiency-level "
            "DISTRIBUTION: it carries headcounts and shares per level, not "
            "the underlying 1.0-6.0 scale scores, so it cannot answer an "
            "'average WIDA score' question. The 2021 COVID file "
            "additionally publishes enrollment-based testing-participation "
            "rates (testing was voluntary/disrupted that year). Coverage: "
            "spring 2017 through spring 2024, state level only — bronze "
            "publishes no district or school breakdown and no demographic "
            "breakouts beyond grade."
        ),
        title="WIDA ACCESS English-Learner Proficiency Results",
        summary=(
            "Statewide WIDA ACCESS English-language-proficiency results for "
            "Georgia English Learners by grade and test domain, 2017-2024."
        ),
        columns=[
            {
                "name": "year",
                "type": "int32",
                "nullable": False,
                "example": 2024,
                "description": (
                    "Spring testing year, the ending calendar year of the "
                    "school year (2024 = school year 2023-2024). Extracted "
                    "from the filename and cross-checked against the "
                    "sheet's title row."
                ),
            },
            {
                "name": "district_code",
                "type": "string",
                "example": None,
                "null_meaning": (
                    "NULL = state-level geography (every row), NOT "
                    "suppression — this topic publishes state aggregates "
                    "only."
                ),
                "description": (
                    "Always NULL: bronze reports statewide totals only, "
                    "with no district breakdown."
                ),
            },
            {
                "name": "school_code",
                "type": "string",
                "example": None,
                "null_meaning": (
                    "NULL = state-level geography (every row), NOT "
                    "suppression — this topic publishes state aggregates "
                    "only."
                ),
                "description": (
                    "Always NULL: bronze reports statewide totals only, "
                    "with no school breakdown."
                ),
            },
            {
                "name": "grade_level",
                "type": "string",
                "nullable": False,
                "example": "05",
                "validValues": sorted({"k"} | {f"{g:02d}" for g in range(1, 13)}),
                "description": (
                    "Canonical grade code (`k`, `01`..`12`). Grade is the "
                    "bronze row axis; there is no `demographic` column "
                    "because the source has no race/gender/economic "
                    "breakouts."
                ),
                "short_description": (
                    "Grade tested, k (kindergarten) through 12."
                ),
            },
            {
                "name": "domain",
                "type": "string",
                "nullable": False,
                "example": "listening",
                "validValues": sorted(set(DOMAIN_MAP.values())),
                "description": (
                    "WIDA test domain: four individually tested language "
                    "domains (`listening`, `speaking`, `reading`, "
                    "`writing`) plus four derived composites "
                    "(`oral_language_composite` = 50%% listening + 50%% "
                    "speaking; `literacy_composite` = 50%% reading + 50%% "
                    "writing; `comprehension_composite` = 70%% reading + "
                    "30%% listening; `overall_score_composite` = 35%% "
                    "reading + 35%% writing + 15%% listening + 15%% "
                    "speaking)."
                ),
                "short_description": (
                    "The skill area scored: listening, speaking, reading, "
                    "writing, or one of four composite scores."
                ),
            },
            {
                "name": "proficiency_level",
                "type": "string",
                "nullable": False,
                "example": "level_4_expanding",
                "validValues": sorted(set(PROFICIENCY_LEVEL_MAP.values())),
                "description": (
                    "WIDA proficiency level, the fixed six-level scale from "
                    "`level_1_entering` (lowest) to `level_6_reaching` "
                    "(highest)."
                ),
                "short_description": (
                    "WIDA proficiency band, level 1 (entering) up to level 6 "
                    "(reaching)."
                ),
            },
            {
                "name": "num_at_proficiency_level",
                "type": "int64",
                "unit": "count",
                "metric_component": "numerator",
                "nullable": False,
                "example": 6475,
                "description": (
                    "Count of ELL students in this grade scoring at this "
                    "proficiency level in this domain. Zero is a real zero "
                    "— the source has no suppression (state aggregates "
                    "exceed any small-cell threshold)."
                ),
            },
            {
                "name": "pct_at_proficiency_level",
                "type": "float64",
                "unit": "proportion",
                "key_metric": True,
                "nullable": False,
                "example": 0.4577,
                "description": (
                    "Share of domain-tested students at this level (0-1 "
                    "scale; bronze publishes 0-100). The denominator is "
                    "`num_tested_in_domain` (students tested in THIS "
                    "domain), NOT `num_tested`. The six level shares sum "
                    "to 1.0 per (year, grade_level, domain)."
                ),
                "short_description": (
                    "Share of students tested in this domain who scored at "
                    "this proficiency level, on a 0-1 scale."
                ),
            },
            {
                "name": "num_tested_in_domain",
                "type": "int64",
                "unit": "count",
                "metric_component": "denominator",
                "nullable": False,
                "example": 14148,
                "description": (
                    "Students tested in this specific domain — the "
                    "denominator of `pct_at_proficiency_level`. 2021 "
                    "(Era 2): taken from the published `Total Tested in "
                    "Domain` column. All other years (Era 1): "
                    "reconstructed as the sum of the six "
                    "`num_at_proficiency_level` counts for the (grade, "
                    "domain) — the source leaves the denominator implicit "
                    "(verified: the six published shares sum to 100.0 "
                    "against exactly this total in every file, and the "
                    "2021 published value equals the level-count sum in "
                    "every cell). Always <= `num_tested`."
                ),
            },
            {
                "name": "num_tested",
                "type": "int64",
                "unit": "count",
                "nullable": False,
                "example": 14148,
                "description": (
                    "Students in this grade tested in at least one domain "
                    "— the grade-wide testing headcount, repeated on every "
                    "(domain, level) row of the grade. Not every such "
                    "student tested in every domain, so per-domain "
                    "`num_tested_in_domain` is <= this value."
                ),
            },
            {
                "name": "enrolled_tested_rate",
                "type": "float64",
                "unit": "proportion",
                "example": 0.9545,
                "null_meaning": (
                    "NULL = metric not published that year (only the 2021 "
                    "COVID file reports it), NOT suppression."
                ),
                "description": (
                    "Share of enrolled ELLs in this grade who tested in at "
                    "least one domain (0-1 scale). Published only for 2021, "
                    "when ACCESS participation was voluntary/disrupted; "
                    "NULL for 2017-2020 and 2022-2024 (the enrolled-ELL "
                    "denominator was never published outside 2021)."
                ),
            },
            {
                "name": "enrolled_tested_in_domain_rate",
                "type": "float64",
                "unit": "proportion",
                "example": 0.9541,
                "null_meaning": (
                    "NULL = metric not published that year (only the 2021 "
                    "COVID file reports it), NOT suppression."
                ),
                "description": (
                    "Share of enrolled ELLs in this grade who tested in "
                    "this domain (0-1 scale) — for composites, in every "
                    "component domain (the source labels these `Tested in "
                    "Both Domains` for the two-domain composites and "
                    "`Tested in All Four Domains` for the overall score). "
                    "Published only for 2021; NULL for all other years."
                ),
            },
        ],
        source="Georgia Insights (GaDOE)",
        source_url=SOURCE_URL,
        update_frequency="annual",
        year_range=year_range,
        partitioned_by=["year"],
        suppressed_to_null=False,
        # 2.0.0: breaking rename of the two participation-rate columns to the
        # canonical vocabulary (enrollment_tested_rate -> enrolled_tested_rate,
        # enrollment_tested_in_domain_rate -> enrolled_tested_in_domain_rate).
        version="2.0.0",
        notes=[
            (
                "State detail level only: `district_code` and `school_code` "
                "are NULL on every row; only states.parquet is emitted per "
                "year. Bronze publishes no district/school breakdown."
            ),
            (
                "No suppression anywhere in the source — a 0 count is a "
                "real zero. NULLs are structural: geography keys (state "
                "level) and the two participation rates outside 2021."
            ),
            (
                "`pct_at_proficiency_level` uses the per-domain denominator "
                "(`num_tested_in_domain`), not `num_tested` — the six level "
                "shares sum to 1.0 per (year, grade_level, domain)."
            ),
            (
                "Era 1 (2017-2020, 2022-2024) leaves the per-domain "
                "denominator implicit; the transform reconstructs it as the "
                "sum of the six level counts. Era 2 (2021) publishes it "
                "explicitly, and the published value equals the level-count "
                "sum in every cell."
            ),
            (
                "2021 composite domain headers carry trailing footnote "
                "letters (`Oral Language CompositeA` .. `Overall Score "
                "CompositeD`) referencing the composite-weighting footnote "
                "rows; both spellings normalize to the same gold domain "
                "codes."
            ),
            (
                "No file exists for 2016 or earlier — Georgia Insights "
                "publishes these state-results workbooks starting with "
                "spring 2017. Unlike GOSA assessment topics, there is no "
                "2020/2021 gap: WIDA ACCESS testing continued through the "
                "pandemic."
            ),
        ],
        quality_checks=[
            {
                "name": "level_counts_sum_to_domain_denominator",
                "description": (
                    "The six per-level `num_at_proficiency_level` counts "
                    "sum exactly to `num_tested_in_domain` for every "
                    "(year, grade_level, domain) group. Era 1 holds by "
                    "construction (the denominator is that sum); Era 2 "
                    "verified against the published 2021 denominators (0 "
                    "mismatches in bronze)."
                ),
                "dimension": "consistency",
                "query": (
                    "SELECT COUNT(*) FROM ("
                    "SELECT year, grade_level, domain "
                    "FROM {object} "
                    "GROUP BY year, grade_level, domain "
                    "HAVING SUM(num_at_proficiency_level) "
                    "<> MAX(num_tested_in_domain)"
                    ")"
                ),
                "mustBe": 0,
            },
            {
                "name": "level_shares_sum_to_one",
                "description": (
                    "The six per-level `pct_at_proficiency_level` shares "
                    "sum to 1.0 within 0.001 for every (year, grade_level, "
                    "domain) group. Verified in bronze: max deviation from "
                    "100.0 across all 8 files is 2e-06 (2e-08 after "
                    "scaling)."
                ),
                "dimension": "consistency",
                "query": (
                    "SELECT COUNT(*) FROM ("
                    "SELECT year, grade_level, domain "
                    "FROM {object} "
                    "GROUP BY year, grade_level, domain "
                    "HAVING ABS(SUM(pct_at_proficiency_level) - 1.0) > 0.001"
                    ")"
                ),
                "mustBe": 0,
            },
            {
                "name": "count_ordering_chain",
                "description": (
                    "num_at_proficiency_level <= num_tested_in_domain <= "
                    "num_tested on every row: a level count cannot exceed "
                    "its domain's tested total, and a domain's tested "
                    "total cannot exceed the students tested in at least "
                    "one domain. Verified on all 8 bronze files: 0 "
                    "violations."
                ),
                "dimension": "consistency",
                "query": (
                    "SELECT COUNT(*) FROM {object} WHERE "
                    "(num_at_proficiency_level IS NOT NULL AND "
                    "num_tested_in_domain IS NOT NULL AND "
                    "num_at_proficiency_level > num_tested_in_domain) OR "
                    "(num_tested_in_domain IS NOT NULL AND "
                    "num_tested IS NOT NULL AND "
                    "num_tested_in_domain > num_tested)"
                ),
                "mustBe": 0,
            },
            {
                "name": "participation_rates_2021_only",
                "description": (
                    "`enrolled_tested_rate` and "
                    "`enrolled_tested_in_domain_rate` are populated on "
                    "every 2021 row and NULL on every other year's rows — "
                    "the enrollment-participation metrics exist only in "
                    "the 2021 COVID file."
                ),
                "dimension": "completeness",
                "query": (
                    "SELECT COUNT(*) FROM {object} WHERE "
                    "((year = 2021) <> (enrolled_tested_rate IS NOT NULL)) "
                    "OR ((year = 2021) <> "
                    "(enrolled_tested_in_domain_rate IS NOT NULL))"
                ),
                "mustBe": 0,
            },
            {
                "name": "geography_always_null",
                "description": (
                    "Structural fact: this is a state-only topic, so "
                    "`district_code` and `school_code` are NULL on every "
                    "row."
                ),
                "dimension": "consistency",
                "query": (
                    "SELECT COUNT(*) FROM {object} WHERE "
                    "district_code IS NOT NULL OR school_code IS NOT NULL"
                ),
                "mustBe": 0,
            },
        ],
    )


if __name__ == "__main__":
    main()
